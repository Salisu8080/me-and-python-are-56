
# Requirements
'''
pip install openpyxl xlwt
Written by Engr. Salisu Zubairu Gaya

'''
import openpyxl
import xlwt
import os

def convert_xlsx_to_xls(input_dir, output_dir):
    try:
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        # Loop through all .xlsx files in the input directory
        for filename in os.listdir(input_dir):
            if filename.endswith('.xlsx'):
                input_file = os.path.join(input_dir, filename)
                output_file = os.path.join(output_dir, filename.replace('.xlsx', '.xls'))

                # Load the .xlsx workbook
                xlsx_workbook = openpyxl.load_workbook(input_file)

                # Create a new .xls workbook
                xls_workbook = xlwt.Workbook()

                # Iterate through all sheets in the .xlsx workbook
                for sheet_name in xlsx_workbook.sheetnames:
                    xlsx_sheet = xlsx_workbook[sheet_name]
                    xls_sheet = xls_workbook.add_sheet(sheet_name[:31])  # Sheet names in .xls are limited to 31 characters

                    # Copy cell values and styles (if needed)
                    for row_index, row in enumerate(xlsx_sheet.iter_rows()):
                        for col_index, cell in enumerate(row):
                            value = cell.value
                            if value is not None:
                                xls_sheet.write(row_index, col_index, value)

                # Save the .xls workbook
                xls_workbook.save(output_file)
                print(f"Converted: {input_file} -> {output_file}")

        print("Conversion completed for all files.")

    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
input_directory = "./"  # Replace with your input directory path
output_directory = "Output_folder"  # Replace with your output directory path
convert_xlsx_to_xls(input_directory, output_directory)
